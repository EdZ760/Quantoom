import serial
import math
import time
import json
import random
import pyperclip
from collections import Counter
from datetime import datetime
from openpyxl import load_workbook
import sys

# --- CONFIGURATION ---
PORT = 'COM4' 
BAUD = 115200
SOURCE_DICT = r'C:\Users\iliut\Desktop\NewFish\Python\words_dictionary2.json'
SHUFFLED_DICT = r'C:\Users\iliut\Desktop\NewFish\Python\words_dictionary3.json'
EXCEL_FILE = r'c:\Users\iliut\Desktop\NewFish\Python\Fruit.xlsx'

def get_bit(ser):
    """Read a single bit from the hardware."""
    while True:
        if ser.in_waiting > 0:
            bit = ser.read(1).decode('utf-8', errors='ignore')
            if bit in ['0', '1']:
                return int(bit)

def get_uint16(ser):
    """Assemble 16 bits into an integer (0-65535)."""
    val = 0
    for i in range(16):
        bit = get_bit(ser)
        val |= (bit << i)
    return val

def get_word_from_file(file_path, line_numbers):
    try:
        with open(file_path, 'r') as f:
            lines = f.readlines()
            words = []
            for n in line_numbers:
                idx = int(n) - 1
                if 0 <= idx < len(lines):
                    line = lines[idx].strip()
                    words.append(line.split()[0] if line.split() else None)
                else:
                    words.append(None)
            return words
    except FileNotFoundError:
        return None

def calculate_metrics(bit_list):
    if not bit_list: return 0, 0
    counts = Counter(bit_list)
    total = len(bit_list)
    entropy = -sum((c/total) * math.log2(c/total) for c in counts.values())
    bias = abs(bit_list.count('1') - bit_list.count('0')) / total * 100
    return entropy, bias

def process_one_shot(ser):
    """Runs exactly once when entropy conditions are met, then exits."""
    print("\n[THRESHOLD MET: PROCESSING]")
    
    # 1. Generate seeds from hardware
    raw_numbers = [get_uint16(ser) for _ in range(5)]
    session_seed = int("".join(map(str, raw_numbers)))
    
    # 2. Shuffle dictionary using the seed
    try:
        with open(SOURCE_DICT, 'r') as f:
            lines = f.readlines()
        random.seed(session_seed)
        random.shuffle(lines)
        with open(SHUFFLED_DICT, 'w') as f:
            f.writelines(lines)
        print("Dictionary shuffled.")
    except Exception as e:
        print(f"File Error: {e}")
        sys.exit()

    # 3. Extract word based on scaled indices
    indices = [(x / 65535.0) * 370101 for x in raw_numbers]
    words = get_word_from_file(SHUFFLED_DICT, indices)
    
    if words and words[0]:
        clean_word = ''.join(c for c in words[0] if c.isalpha())
        print(f"WORD GENERATED: {clean_word}")
        pyperclip.copy(clean_word)
        
        # 4. Log to Excel
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            row = 1
            while ws.cell(row=row, column=1).value:
                row += 1
            ws.cell(row=row, column=1).value = clean_word
            ws.cell(row=row, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            wb.save(EXCEL_FILE)
            print(f"Logged to row {row}.")
        except Exception as e:
            print(f"Excel Error: {e}")

    # 5. Geolocation output
    coords = [f"{(raw_numbers[i]/65535.0)*180-90:.6f}, {(raw_numbers[i+1]/65535.0)*360-180:.6f}"
              for i in range(0, len(raw_numbers)-1, 2)]
    print("Coordinates:\n" + "\n".join(coords))
    
    # 6. EXIT
    print("\n[SUCCESS] Shutting down.")
    ser.close()
    sys.exit()

# --- MAIN MONITORING LOOP ---
try:
    ser = serial.Serial(PORT, BAUD, timeout=1)
    print(f"Monitoring {PORT}. Waiting for Entropy > 0.99 and Bias < 1%...")
    
    while True:
        raw = ser.read(1000).decode('utf-8', errors='ignore')
        bits = [c for c in raw if c in '01']
        
        if len(bits) > 100:
            entropy, bias = calculate_metrics(bits)
            print(f"Entropy: {entropy:.4f} | Bias: {bias:.2f}%", end='\r')
            
            if entropy >= 1.0 and bias < 1.0:
                process_one_shot(ser)
                
except KeyboardInterrupt:
    if 'ser' in locals() and ser.is_open:
        ser.close()
    print("\nTerminated.")